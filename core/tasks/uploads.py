# core/tasks/uploads.py
"""
Async upload processing tasks.

Design summary (see accompanying explanation for full rationale):

    Storage (Supabase) → temp local file → openpyxl read-only iteration
        → per-chunk vectorized transform (pandas, no iterrows)
        → CSV buffer → COPY FROM STDIN into staging table
        → (repeat per chunk, updating UploadJob.processed_rows)
        → allocate_upload_job_ids (persistent global identity registries)
        → flush_*_upload_job (transactional and job-scoped)
        → UploadJob marked success/partial/failed

Every business rule from the old synchronous pipeline is preserved:
    - mapping integrity validated the same way (via the same Excel mapper
      classes' validate_mapping_integrity)
    - column names actually present in the file are captured into
      column_mapping JSON, same as before
    - rows missing a required internal id are skipped (partial success),
      same as before
    - product attribute empty-cell → NULL handling preserved
    - order_date parsing via FlexibleDateParser preserved
    - coupon duplicate-active-coupon guard preserved
    - global unique id allocation preserved, silently, with no user-facing
      change in behavior — only where and how it executes has changed
"""

import io

import openpyxl
import pandas as pd
from celery import shared_task

from django.core.files.base import ContentFile
from django.db import connection, transaction

from core.models import (
    Coupon,
    CouponFileUpload,
    CustomerFileUpload,
    ErrorLog,
    ProductFileUpload,
    UploadJob,
)
from core.services.storage import delete_from_storage, download_to_tempfile
from core.utils.date_parser import FlexibleDateParser
from core.utils.excel_mapper import (
    CouponExcelMapper,
    CustomerExcelMapper,
    ProductExcelMapper,
)

CHUNK_SIZE = 50_000  # overridden by settings.UPLOAD_CHUNK_SIZE where imported


def _chunk_size():
    from django.conf import settings

    return getattr(settings, "UPLOAD_CHUNK_SIZE", CHUNK_SIZE)


def _statement_timeout_ms() -> int:
    from django.conf import settings

    return getattr(settings, "UPLOAD_DB_STATEMENT_TIMEOUT_MS", 7_200_000)


def _claim_job(job_id: str, upload_type: str, task_id: str) -> UploadJob | None:
    """Atomically claim a job while making Celery redelivery idempotent."""
    with transaction.atomic():
        job = (
            UploadJob.objects.select_for_update()
            .select_related("tenant")
            .get(id=job_id)
        )
        if job.upload_type != upload_type:
            raise ValueError(f"Upload job {job_id} is not a {upload_type} job.")
        if job.status in (UploadJob.Status.SUCCESS, UploadJob.Status.PARTIAL):
            return None
        if (
            job.status == UploadJob.Status.PROCESSING
            and job.celery_task_id
            and job.celery_task_id != task_id
        ):
            return None
        job.status = UploadJob.Status.PROCESSING
        job.celery_task_id = task_id
        job.error_type = None
        job.message = ""
        job.processed_rows = 0
        job.rows_saved = 0
        job.save(
            update_fields=[
                "status",
                "celery_task_id",
                "error_type",
                "message",
                "processed_rows",
                "rows_saved",
                "updated_at",
            ]
        )
        return job


def _clear_job_staging(job: UploadJob) -> None:
    """Remove only this job's prior partial COPY output before a retry."""
    table = {
        UploadJob.UploadType.CUSTOMERS: "users_unnormalized_data_staging",
        UploadJob.UploadType.PRODUCTS: "products_unnormalized_data_staging",
    }.get(job.upload_type)
    if table is None:
        return
    with connection.cursor() as cursor:
        cursor.execute(f"DELETE FROM {table} WHERE upload_job_id = %s", [job.id])


def _allocate_flush_and_finalize(
    job: UploadJob,
    flush_function: str,
    final_status: str,
    success_message_template: str,
) -> int:
    """Allocate, flush, and mark terminal success in one transaction.

    Keeping the UploadJob update in this transaction closes the crash window
    where business rows had committed but Celery still saw the job as
    PROCESSING and could replay the entire file.
    """
    with transaction.atomic():
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT set_config('statement_timeout', %s, true)",
                [str(_statement_timeout_ms())],
            )
            cursor.execute("SELECT allocate_upload_job_ids(%s)", [job.id])
            cursor.execute(f"SELECT {flush_function}(%s)", [job.id])
            rows_saved = cursor.fetchone()[0] or 0

        job.rows_saved = rows_saved
        job.status = final_status
        job.message = success_message_template.format(rows_saved=rows_saved)
        job.save(update_fields=["rows_saved", "status", "message", "updated_at"])
        return rows_saved


def _inspect_workbook(path: str) -> tuple[int, list[str]]:
    """
    Row count without loading the whole sheet into memory. openpyxl in
    read-only mode streams rows rather than materializing the full workbook,
    which is what makes this safe to run on multi-million-row files where
    pd.read_excel(path) up front would already be the first bottleneck.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header = next(ws.iter_rows(values_only=True), ())
    total = ws.max_row - 1  # minus header row
    wb.close()
    return max(total, 0), _column_headers(header)


def _iter_chunks(path: str, chunk_size: int):
    """
    Yields (header_row, list_of_data_rows) chunks by streaming the worksheet
    with openpyxl's read-only iterator, instead of pandas loading the entire
    file into a single DataFrame up front. This bounds peak memory usage to
    one chunk's worth of rows regardless of total file size.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    row_iter = ws.iter_rows(values_only=True)
    header = next(row_iter)

    buffer = []
    for row in row_iter:
        buffer.append(row)
        if len(buffer) >= chunk_size:
            yield header, buffer
            buffer = []
    if buffer:
        yield header, buffer
    wb.close()


def _column_headers(header) -> list[str]:
    return ["" if value is None else str(value) for value in header]


def _copy_buffer_into(table_and_columns_sql: str, buf: io.StringIO):
    """
    Executes a COPY FROM STDIN using psycopg3's actual API.

    psycopg3 does NOT have cursor.copy_expert() (that is a psycopg2-only
    method). psycopg3's equivalent is cursor.copy(sql) used as a context
    manager, into which you .write() chunks of data. This project's
    Django DATABASES connection uses psycopg (v3) -- confirmed via
    `connection.Database` -- so this is the correct call for this
    environment. If this project ever moves to psycopg2, this is the
    only function that needs to change.
    """
    with connection.cursor() as cursor:
        raw_cursor = cursor.cursor  # underlying psycopg3 cursor object
        with raw_cursor.copy(table_and_columns_sql) as copy:
            copy.write(buf.read())


def _mark_failed(job: UploadJob, error_type: str, message: str):
    job.status = UploadJob.Status.FAILED
    job.error_type = error_type
    job.message = message
    job.save(update_fields=["status", "error_type", "message", "updated_at"])

    # Mirrors the existing ErrorLog pattern used elsewhere in the project
    # (core.models.ErrorLog) so upload failures show up in the same place
    # ops/admin already look for campaign/SMS/trigger errors, instead of
    # being visible only via the UploadJob row itself.
    ErrorLog.objects.create(
        tenant=job.tenant,
        source="other",
        severity="error",
        error_code=error_type,
        message=message,
        context={
            "upload_job_id": str(job.id),
            "upload_type": job.upload_type,
            "storage_key": job.storage_key,
        },
    )


def _attach_uploaded_file(
    upload_record, field_name: str, local_path: str, original_filename: str
):
    """
    Saves the Storage-downloaded bytes into the FileField on the audit
    record (CustomerFileUpload.customers_file, etc.), preserving the exact
    old behavior of that field being a real, retrievable file rather than
    left empty. These FileFields are NOT null=True in core/models.py, so
    skipping this save would either raise on record creation or leave an
    unusable empty file reference.
    """
    with open(local_path, "rb") as f:
        getattr(upload_record, field_name).save(
            original_filename, ContentFile(f.read()), save=True
        )


# ─────────────────────────────────────────────────────────────────────────────
# Customers
# ─────────────────────────────────────────────────────────────────────────────


@shared_task(bind=True, name="core.tasks.uploads.process_customers_upload")
def process_customers_upload(self, job_id: str):
    job = _claim_job(job_id, UploadJob.UploadType.CUSTOMERS, self.request.id)
    if job is None:
        return
    tenant = job.tenant
    mapping = job.mapping

    is_valid, error_msg = CustomerExcelMapper.validate_mapping_integrity(mapping)
    if not is_valid:
        _mark_failed(
            job, "mapping_error", f"خطا در نگاشت ستون‌های فایل مشتریان: {error_msg}"
        )
        return

    total_saved = 0
    total_processed = 0
    actual_column_mapping = None
    upload_record = None
    succeeded = False

    try:
        _clear_job_staging(job)
        with download_to_tempfile(job.storage_key) as local_path:
            # Create the audit record here, inside the download context, so
            # the FileField can be populated from the same local copy in one
            # pass rather than downloading twice. Required because
            # CustomerFileUpload.customers_file is a real FileField with no
            # null=True in core/models.py -- it needs actual bytes.
            upload_record, created = CustomerFileUpload.objects.get_or_create(
                upload_job=job,
                defaults={"tenant": tenant, "customers_mapping": mapping},
            )
            if created or not upload_record.customers_file:
                _attach_uploaded_file(
                    upload_record, "customers_file", local_path, job.original_filename
                )

            job.total_rows, job.column_headers = _inspect_workbook(local_path)
            job.save(update_fields=["total_rows", "column_headers", "updated_at"])

            for header, rows in _iter_chunks(local_path, _chunk_size()):
                if actual_column_mapping is None:
                    actual_column_mapping = {
                        field: header[idx] for field, idx in mapping.items()
                    }

                chunk_df = pd.DataFrame(rows, columns=header)
                saved_in_chunk, processed_in_chunk = _copy_customers_chunk(
                    upload_job_id=str(job.id),
                    tenant_id=tenant.id,
                    chunk_df=chunk_df,
                    mapping=mapping,
                    actual_column_mapping=actual_column_mapping,
                    upload_created_at=str(upload_record.created_at),
                )
                total_saved += saved_in_chunk
                total_processed += processed_in_chunk

                job.processed_rows = total_processed
                job.save(update_fields=["processed_rows", "updated_at"])

        if total_saved == 0:
            _mark_failed(
                job, "file_error", "هیچ رکورد معتبری در فایل مشتریان یافت نشد."
            )
            return

        final_status = (
            UploadJob.Status.SUCCESS
            if total_processed == job.total_rows
            else UploadJob.Status.PARTIAL
        )
        _allocate_flush_and_finalize(
            job,
            "flush_customers_upload_job",
            final_status,
            "{rows_saved} رکورد مشتریان با موفقیت در دیتابیس ذخیره شد.",
        )
        succeeded = True

    except Exception as e:
        _mark_failed(
            job, "pipeline_error", f"خطای غیرمنتظره در پردازش فایل مشتریان: {str(e)}"
        )
    finally:
        if succeeded:
            delete_from_storage(job.storage_key)


def _copy_customers_chunk(
    upload_job_id,
    tenant_id,
    chunk_df,
    mapping,
    actual_column_mapping,
    upload_created_at,
):
    """
    Vectorized (no iterrows) transform of one chunk, then COPY into
    users_unnormalized_data_staging. Mirrors exactly what the old per-row
    loop did field by field, just expressed as column operations.

    Returns (rows_saved_in_chunk, rows_processed_in_chunk).
    """
    df = pd.DataFrame()
    df["internal_user_id"] = chunk_df.iloc[:, mapping["internal_id"]].astype(str)
    df["internal_order_id"] = chunk_df.iloc[:, mapping["internal_order_id"]].astype(str)
    df["internal_product_id"] = chunk_df.iloc[:, mapping["internal_product_id"]].astype(
        str
    )
    df["first_name"] = (
        chunk_df.iloc[:, mapping["first_name"]].astype(str).str.slice(0, 200)
    )
    df["last_name"] = (
        chunk_df.iloc[:, mapping["last_name"]].astype(str).replace("", None)
    )
    df["gender"] = chunk_df.iloc[:, mapping["gender"]].astype(str).replace("", None)
    df["phone_number"] = (
        chunk_df.iloc[:, mapping["phone_number"]]
        .astype(str)
        .str.slice(0, 20)
        .replace("", None)
    )
    df["order_date"] = FlexibleDateParser.parse_series_to_dates(
        chunk_df.iloc[:, mapping["order_date"]]
    )
    df["then_product_price"] = pd.to_numeric(
        chunk_df.iloc[:, mapping["then_product_price"]], errors="coerce"
    ).fillna(0.0)
    df["quantity"] = (
        pd.to_numeric(chunk_df.iloc[:, mapping["quantity"]], errors="coerce")
        .fillna(0)
        .astype(int)
    )

    rows_processed = len(df)

    # Same required-field skip rule as before: internal_id, internal_order_id,
    # internal_product_id must all be non-empty.
    valid_mask = (
        (df["internal_user_id"] != "")
        & (df["internal_user_id"] != "None")
        & (df["internal_order_id"] != "")
        & (df["internal_order_id"] != "None")
        & (df["internal_product_id"] != "")
        & (df["internal_product_id"] != "None")
    )
    df = df[valid_mask]

    if df.empty:
        return 0, rows_processed

    column_mapping_json = _json_for_copy(
        {
            "customers_file_mapping": actual_column_mapping,
            "customers_index_mapping": mapping,
            "uploaded_at": upload_created_at,
        }
    )

    out = pd.DataFrame(
        {
            "upload_job_id": upload_job_id,
            "tenant_id": tenant_id,
            "internal_user_id": df["internal_user_id"],
            "first_name": df["first_name"],
            "last_name": df["last_name"],
            "gender": df["gender"],
            "phone_number": df["phone_number"],
            "internal_order_id": df["internal_order_id"],
            "order_date": df["order_date"],
            "internal_product_id": df["internal_product_id"],
            "then_product_price": df["then_product_price"],
            "quantity": df["quantity"],
            "column_mapping": column_mapping_json,
        }
    )

    buf = io.StringIO()
    # Fully vectorized CSV serialization -- no Python-level per-row loop at
    # all. This is the direct fix for the iterrows() bottleneck the original
    # pipeline had, carried all the way through to the COPY boundary.
    out.to_csv(buf, index=False, header=False, na_rep="")
    buf.seek(0)

    _copy_buffer_into(
        """
        COPY users_unnormalized_data_staging (
            upload_job_id, tenant_id, internal_user_id, first_name, last_name,
            gender, phone_number, internal_order_id, order_date,
            internal_product_id, then_product_price, quantity,
            column_mapping
        ) FROM STDIN WITH (FORMAT csv, NULL '')
        """,
        buf,
    )

    return len(df), rows_processed


def _json_for_copy(d: dict) -> str:
    import json

    # COPY CSV mode needs the JSON string escaped so embedded commas/quotes
    # don't break CSV parsing — csv.writer handles quoting for us as long as
    # we pass a plain string field, so just serialize normally here.
    return json.dumps(d, ensure_ascii=False)


# ─────────────────────────────────────────────────────────────────────────────
# Products
# ─────────────────────────────────────────────────────────────────────────────


@shared_task(bind=True, name="core.tasks.uploads.process_products_upload")
def process_products_upload(self, job_id: str):
    job = _claim_job(job_id, UploadJob.UploadType.PRODUCTS, self.request.id)
    if job is None:
        return
    tenant = job.tenant
    mapping = job.mapping

    is_valid, error_msg = ProductExcelMapper.validate_mapping_integrity(mapping)
    if not is_valid:
        _mark_failed(
            job, "mapping_error", f"خطا در نگاشت ستون‌های فایل محصولات: {error_msg}"
        )
        return

    total_saved = 0
    total_processed = 0
    actual_column_mapping = None
    upload_record = None
    succeeded = False

    try:
        _clear_job_staging(job)
        with download_to_tempfile(job.storage_key) as local_path:
            upload_record, created = ProductFileUpload.objects.get_or_create(
                upload_job=job,
                defaults={"tenant": tenant, "products_mapping": mapping},
            )
            if created or not upload_record.products_file:
                _attach_uploaded_file(
                    upload_record, "products_file", local_path, job.original_filename
                )

            job.total_rows, job.column_headers = _inspect_workbook(local_path)
            job.save(update_fields=["total_rows", "column_headers", "updated_at"])

            for header, rows in _iter_chunks(local_path, _chunk_size()):
                if actual_column_mapping is None:
                    actual_column_mapping = {
                        field: header[idx] for field, idx in mapping.items()
                    }

                chunk_df = pd.DataFrame(rows, columns=header)
                saved_in_chunk, processed_in_chunk = _copy_products_chunk(
                    upload_job_id=str(job.id),
                    tenant_id=tenant.id,
                    chunk_df=chunk_df,
                    mapping=mapping,
                    actual_column_mapping=actual_column_mapping,
                    upload_created_at=str(upload_record.created_at),
                )
                total_saved += saved_in_chunk
                total_processed += processed_in_chunk

                job.processed_rows = total_processed
                job.save(update_fields=["processed_rows", "updated_at"])

        if total_saved == 0:
            _mark_failed(
                job, "file_error", "هیچ رکورد معتبری در فایل محصولات یافت نشد."
            )
            return

        final_status = (
            UploadJob.Status.SUCCESS
            if total_processed == job.total_rows
            else UploadJob.Status.PARTIAL
        )
        _allocate_flush_and_finalize(
            job,
            "flush_products_upload_job",
            final_status,
            "{rows_saved} رکورد محصولات با موفقیت در دیتابیس ذخیره شد.",
        )
        succeeded = True

    except Exception as e:
        _mark_failed(
            job, "pipeline_error", f"خطای غیرمنتظره در پردازش فایل محصولات: {str(e)}"
        )
    finally:
        if succeeded:
            delete_from_storage(job.storage_key)


def _copy_products_chunk(
    upload_job_id,
    tenant_id,
    chunk_df,
    mapping,
    actual_column_mapping,
    upload_created_at,
):
    df = pd.DataFrame()
    df["internal_product_id"] = chunk_df.iloc[:, mapping["internal_product_id"]].astype(
        str
    )
    df["product_name"] = (
        chunk_df.iloc[:, mapping["product_name"]].astype(str).str.slice(0, 255)
    )
    df["category"] = chunk_df.iloc[:, mapping["category"]].astype(str).str.slice(0, 100)
    df["current_product_price"] = pd.to_numeric(
        chunk_df.iloc[:, mapping["current_product_price"]], errors="coerce"
    ).fillna(0.0)
    df["product_link"] = (
        chunk_df.iloc[:, mapping["product_link"]].astype(str).str.slice(0, 2000)
    )

    # Attribute fields: empty cell -> NULL, preserved exactly as before.
    # Using pandas' own NaN-awareness rather than Python-side per-row check.
    first_attr_raw = chunk_df.iloc[:, mapping["first_product_attribute"]]
    second_attr_raw = chunk_df.iloc[:, mapping["second_product_attribute"]]
    df["first_product_attribute"] = (
        first_attr_raw.astype(str)
        .str.strip()
        .replace({"": None, "nan": None, "None": None})
    )
    df["second_product_attribute"] = (
        second_attr_raw.astype(str)
        .str.strip()
        .replace({"": None, "nan": None, "None": None})
    )

    rows_processed = len(df)

    valid_mask = (df["internal_product_id"] != "") & (
        df["internal_product_id"] != "None"
    )
    df = df[valid_mask]

    if df.empty:
        return 0, rows_processed

    column_mapping_json = _json_for_copy(
        {
            "products_file_mapping": actual_column_mapping,
            "products_index_mapping": mapping,
            "uploaded_at": upload_created_at,
        }
    )

    out = pd.DataFrame(
        {
            "upload_job_id": upload_job_id,
            "tenant_id": tenant_id,
            "internal_product_id": df["internal_product_id"],
            "product_name": df["product_name"],
            "category": df["category"],
            "current_product_price": df["current_product_price"],
            "product_link": df["product_link"],
            "first_product_attribute": df["first_product_attribute"],
            "second_product_attribute": df["second_product_attribute"],
            "column_mapping": column_mapping_json,
        }
    )

    buf = io.StringIO()
    out.to_csv(buf, index=False, header=False, na_rep="")
    buf.seek(0)

    _copy_buffer_into(
        """
        COPY products_unnormalized_data_staging (
            upload_job_id, tenant_id, internal_product_id,
            product_name, product_category,
            current_product_price, product_link,
            first_product_attribute, second_product_attribute,
            column_mapping
        ) FROM STDIN WITH (FORMAT csv, NULL '')
        """,
        buf,
    )

    return len(df), rows_processed


# ─────────────────────────────────────────────────────────────────────────────
# Coupons — kept simple/synchronous-style logic but still off the request
# thread. Coupons files are not expected to be multi-million-row, but the
# job/status pattern is used uniformly across all three upload types for a
# consistent frontend integration.
# ─────────────────────────────────────────────────────────────────────────────


@shared_task(bind=True, name="core.tasks.uploads.process_coupons_upload")
def process_coupons_upload(self, job_id: str):
    job = UploadJob.objects.select_related("tenant").get(id=job_id)
    tenant = job.tenant
    mapping = job.mapping

    is_valid, error_msg = CouponExcelMapper.validate_mapping_integrity(mapping)
    if not is_valid:
        _mark_failed(
            job, "mapping_error", f"خطا در نگاشت ستون‌های فایل کوپن: {error_msg}"
        )
        return

    # Duplicate coupon guard — unchanged business rule from the sync pipeline.
    if Coupon.objects.filter(tenant=tenant, status="available").exists():
        _mark_failed(
            job,
            "duplicate_coupon_error",
            "شما هنوز کوپن‌های استفاده نشده دارید. تا زمانی که همه کوپن‌ها استفاده "
            "نشده‌اند، امکان آپلود فایل کوپن جدید وجود ندارد.",
        )
        return

    job.status = UploadJob.Status.PROCESSING
    job.celery_task_id = self.request.id
    job.save(update_fields=["status", "celery_task_id", "updated_at"])

    total_saved = 0
    total_processed = 0

    try:
        with download_to_tempfile(job.storage_key) as local_path:
            coupon_upload_record = CouponFileUpload.objects.create(
                tenant=tenant,
                coupons_mapping=mapping,
                upload_job=job,
            )
            _attach_uploaded_file(
                coupon_upload_record, "coupons_file", local_path, job.original_filename
            )

            job.total_rows, job.column_headers = _inspect_workbook(local_path)
            job.save(update_fields=["total_rows", "column_headers", "updated_at"])

            for header, rows in _iter_chunks(local_path, _chunk_size()):
                chunk_df = pd.DataFrame(rows, columns=header)
                codes = chunk_df.iloc[:, mapping["coupon_code"]].astype(str).str.strip()
                discounts = pd.to_numeric(
                    chunk_df.iloc[:, mapping["discount_percentage"]], errors="coerce"
                ).fillna(0)

                total_processed += len(chunk_df)
                mask = codes != ""
                codes, discounts = codes[mask], discounts[mask]

                coupon_objects = [
                    Coupon(
                        tenant=tenant,
                        coupon_code=code,
                        discount_percentage=discount,
                        status="available",
                    )
                    for code, discount in zip(codes, discounts)
                ]
                if coupon_objects:
                    with transaction.atomic():
                        Coupon.objects.bulk_create(
                            coupon_objects, batch_size=5000, ignore_conflicts=True
                        )
                    total_saved += len(coupon_objects)

                job.processed_rows = total_processed
                job.save(update_fields=["processed_rows", "updated_at"])

        if total_saved == 0:
            _mark_failed(job, "file_error", "هیچ کوپن معتبری در فایل یافت نشد.")
            return

        job.rows_saved = total_saved
        job.status = (
            UploadJob.Status.SUCCESS
            if total_processed == job.total_rows
            else UploadJob.Status.PARTIAL
        )
        job.message = f"{total_saved} کوپن با موفقیت در دیتابیس ذخیره شد."
        job.save(update_fields=["rows_saved", "status", "message", "updated_at"])

    except Exception as e:
        _mark_failed(
            job, "pipeline_error", f"خطای غیرمنتظره در ذخیره‌سازی کوپن‌ها: {str(e)}"
        )
    finally:
        delete_from_storage(job.storage_key)
